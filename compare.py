from openpyxl import load_workbook

wb = load_workbook("compare.xlsx")
ws = wb.active
colA = ws['A']
colC = ws['C']
registered = []
for cell in colC:
    if cell.value is None:
        break
    registered.append(cell.value)

for cell in colA:
    if cell.value is None:
        break
    if cell.value not in registered:
        print(cell.value)

# print(registered)
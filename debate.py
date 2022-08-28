"""
Program to help in Aarohan Debate
This program has been made to make the Aarohan Debate go smoother.

The program checks whether a participant has registered for the event,
checks their status, logs the time the participant leaves for the venue 
and what room they go to. It also allows for changes to be 
done in the case of any mistakes in user entry. 

The program integrates with Excel to save data as it happens. It logs
the timings of arrivals and departures from the mini audi.

Idea by : Venkataram S
Author : Sam Alex Koshy
Starting date : 2022/08/22 

"""

# Importing libraries.
import sys
import os
from datetime import datetime
from time import sleep

import re
# import json
# import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook

class Debate():
    """An object representing a workbook to work on."""

    def __init__(self):
        """Initialize the program with important details."""
        # Load workbook.
        try:
            self.workbook = load_workbook("debate.xlsx")
            print("Workbook Loaded")
        except:
            print("Workbook could not be loaded")
        self.worksheet = self.workbook.active
        self.cmds = "Commands are: check <no>, add <no>, send [no], "
        self.cmds += "name <name>, list, change, save, cmds"
        self.col_names = self.worksheet['A']
        self.col_numbers = self.worksheet['B']

    def main(self):
        """The main functioning of the code. Keeps the code running."""
        print(self.cmds)
        # Keep code running until cmd quit entered.
        while(True):
            cmd = input("\nEnter command : ")
            if cmd.lower().startswith("quit"):
                self._save_workbook(cmd, True)
            elif cmd.lower().startswith("check"):
                self._check_reg_no(cmd)
            elif cmd.lower().startswith("add"):
                self._add_participant(cmd)
            elif cmd.lower().startswith("send"):
                self._send_participant(cmd)
            elif cmd.lower().startswith("name"):
                self._check_name(cmd.lower())
            elif cmd.startswith("change"):
                self._change_participant_details()
            elif cmd.startswith("cmd"):
                print(self.cmds)
            elif cmd.startswith("save"):
                self._save_workbook(cmd, False)

    def _check_reg_no(self, word:str):
        """Check the status of a participant on typing their register number.

        If argument given is not a valid number, the function does not execute.

        This function prints the details of the participant if they 
        are registered, and their status for the debate. If they are not 
        a participant, prompt is given to check their name in the list.

        Parameters
        ---------
        - word `str`:
            The string that contains the register number of the participant, 
            along with the command word.
        """
        try:
            reg_no = int(re.split("check ", word)[1])
        except:
            return

        # Iterate through the column to find the register number
        for cell in self.col_numbers:
            if cell.value == reg_no:
                # If found, print details.
                row:int = cell.row
                msg = ""
                for value in self.worksheet[row]:
                    msg += f"{value.value}, "
                print(msg)

                # Print status of participant.
                if self.worksheet.cell(row=row, column=4).value!=None:
                    print("Already sent the participant")
                elif self.worksheet.cell(row=row, column=3).value!=None:
                    print("Participant is waiting in mini audi.")
                else:
                    print("Participant has not arrived yet.")
                return

        # If register number not found, prompt for name.
        print(reg_no, "not found. Check if their name is there with name <name>.")
        name = input("Enter name : ")
        name = "name " + name
        self._check_name(name)

    def _check_name(self, word:str):
        """Check if the name exists in the list. If it exists, display their 
        details including row number.

        This function prints the details of the participant if their name is found.

        Parameters
        ----------
        - word `str`:
            The string that contains the name of the participant, 
            along with the command word.
        """
        try:
            name = re.split("name ", word)[1]
        except:
            return
        found = False
        deleted = False

        # Iterate through the column to find the name.
        for cell in self.col_names:
            if cell.value is None:
                self.worksheet.delete_rows(cell.row, 1)
                deleted = True
            elif (re.search(cell.value.lower(), name) 
                    or re.search(name, cell.value.lower())):
                # Print the details of the user if found.
                row = cell.row
                msg = f"{row}. "
                for value in self.worksheet[row]:
                    msg += f"{value.value}, "
                print(msg)
                found = True
        if deleted:
            print("Rows were deleted, try again.")
        elif not found:
            print(name, "not found.")

    def _add_participant(self, word:str):
        """Logs the time when the participant enters. 

        This function prints the details of the user, including whether they
        have already been added or if they have been sent.

        Parameters
        ----------
        - word `str`:
            The string that contains the register number of the participant, 
            along with the command word. 
        """
        try:
            reg_no = int(re.split("add", word)[1])
        except:
            return
        found = 0

        # Iterate through the column to find the register number
        for cell in self.col_numbers:
            if cell.value == reg_no:
                row = cell.row
                if self.worksheet.cell(row=row, column=4).value!=None:
                    print("Already sent the participant")
                elif self.worksheet.cell(row=row, column=3).value!=None:
                    print("Already logged in")
                else:
                    self.worksheet.cell(row=row, column=3, value=datetime.now())
                    for value in self.worksheet[cell.row]:
                        print(value.value)
                found = 1
        # If participant is not found, send message.
        if found == 0:
            print(reg_no, "not found.")
        self._auto_save_workbook()

    def _send_participant(self, word:str):
        """Log the time the participant is sent to their panel room, along with
        the room they are sent to.

        This function prints the details of the participant after logging their
        details.

        Parameters
        ----------
        - word `str`:
            The string that contains the register number of the participant, 
            along with the command word. 
        """
        try:
            values = re.findall("\d+", word)
            reg_no = int(values[0])
            panel_room = int(values[1])
        except:
            return

        # Iterate through the column to find the register number
        for cell in self.col_numbers:
            if cell.value == reg_no:
                self.worksheet.cell(row=cell.row, column=4, value=datetime.now())
                self.worksheet.cell(row=cell.row, column=5, value=panel_room)
                for value in self.worksheet[cell.row]:
                    print(value.value)
                print("Sent", reg_no)
        self._auto_save_workbook()

    def _change_participant_details(self):
        """Change the details of any participant.

        The name, register number or class of a participant can be changed.
        """
        # Accept location of the change.
        row = int(input("Enter row address : "))
        col = int(input("Enter column address : "))
        # Validate the location, then make the change
        if col not in [1, 2, 5]:
            print("Invalid column number. Aborting change")
            return
        elif col == 1:
            val = input("Enter name : ")
        elif col == 2:
            val = int(input("Enter reg no : "))
        elif col == 5:
            val = input("Enter class : ")
        self.worksheet.cell(row=row, column=col, value=val)
        msg = f"{row}. "
        for value in self.worksheet[row]:
            msg += f"{value.value}, "
        print(msg)
        self._auto_save_workbook()

    def _save_workbook(self, word:str, close:bool):
        """Save the workbook on demand.

        This function closes all excel windows to ensure that the workbook
        object gets saved to a file.

        Parameters
        ----------
        - word `str`:
            The command whether to open the sheet or not after saving.
        - close `bool`:
            Whether the program needs to be closed after saving.
        """
        # Close all files to ensure a save.
        try:
            os.system("taskkill/im EXCEL.EXE ")
        except:
            print("Already closed")
        # Give some time to ensure all sheets are closed.
        if close:
            sleep(10)
        if re.search("open", word):
            open = True
        else:
            open = False
        # Try to save file.
        try:
            self.workbook.save("debate.xlsx")
        except PermissionError:
            print("Failed to save")
        else:
            print("Excel sheet saved")
            if open:
                os.system("start EXCEL.EXE debate.xlsx")
            # Close the program.
            if close:
                sys.exit()

    def _auto_save_workbook(self):
        """Save the workbook when needed."""
        try:
            self.workbook.save("debate.xlsx")
        except PermissionError:
            print("Failed to save")


if __name__ == "__main__":
    d = Debate()
    d.main()
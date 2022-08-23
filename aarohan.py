import sys
import os
from datetime import datetime
from time import sleep

import openpyxl
from openpyxl import load_workbook

def main():
    wb = load_workbook("E:/quiz.xlsx")
    os.system("start EXCEL.EXE E:/quiz.xlsx")
    while(True):
        cmd = input("\nEnter command:")
        if cmd.lower() == "quit":
            close_window(wb)
        elif cmd.lower()[0].isdigit():
            check_num(wb, cmd)
        elif cmd.lower() == "new":
            new_row(wb)
        elif cmd.lower() == "name":
            check_name(wb)

def check_num(wb, reg_no):
    # reg_no = int(input("Enter number:"))
    reg_no = int(reg_no)
    ws = wb.active
    colB = ws['B']
    for cell in colB:
        if cell.value == reg_no:
            row:int = cell.row
            if ws.cell(row=row, column=3).value==None:
                print(row)
                print(ws.cell(row=row, column=1).value)
                ws.cell(row=row, column=3).value = datetime.now()
                try:
                    wb.save("E:/quiz.xlsx")
                except PermissionError:
                    print("Could not save")
            else:
                print("Already logged in")
            return
    colE = ws['E']
    for cell in colE:
        if cell.value == reg_no:
            row:int = cell.row
            if ws.cell(row=row, column=6).value==None:
                print(row)
                print(ws.cell(row=row, column=4).value)
                ws.cell(row=row, column=6).value = datetime.now()
                try:
                    wb.save("E:/quiz.xlsx")
                except PermissionError:
                    print("Could not save")
            else:
                print("Already logged in")
            return
    print("Not found")

def new_row(wb):
    ws = wb.active
    name = input("Name: ")
    reg_no = int(input("Reg No: "))
    name_2 = input("Name: ")
    reg_no_2 = int(input("Reg No: "))
    time = datetime.now()
    ws.append({
        1: name,
        2: reg_no,
        3: time,
        4: name_2, 
        5: reg_no_2,
        6: time
    })
    print("Added at", time)
    try:
        wb.save("E:/quiz.xlsx")
    except PermissionError:
        print("Could not save")

def check_name(wb):
    pass

def close_window(wb):
    os.system("taskkill/im EXCEL.EXE ")
    sleep(10)
    try:
        wb.save("E:/quiz.xlsx")
        print("File saved")
    except PermissionError:
        print("Failed to save")
    sys.exit()

if __name__ == "__main__":
    main()